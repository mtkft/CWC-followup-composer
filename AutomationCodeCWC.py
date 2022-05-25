# -*- coding: utf-8 -*-
"""
Created on Thu May 19 23:09:49 2022

@author: thepr
"""
import docx
import openpyxl
def getdata(filepath, lineNumber, sheetnum = 0):
    workbook = openpyxl.load_workbook(filename = filepath, data_only=True)
    sheetToRead = workbook.worksheets[sheetnum]
    titleGenerator = sheetToRead.iter_rows(1, 1, values_only = True)
    keys = next(titleGenerator)
    keyGenerator = sheetToRead.iter_rows(lineNumber, lineNumber, values_only = True)
    values = next(keyGenerator)
    output = dict()
    for i in range(len(keys)):
        output[keys[i]] = values[i]
        
    return output
def createparagraph1(filepath, data, newFilepath):
    document = docx.Document(filepath)
    text = f"{data['Resident Name']}\n{data['Building/House Number']} "
    document.add_paragraph()
    return None
