# -*- coding: utf-8 -*-
"""
Created on Fri May 20 20:55:46 2022

@author: thepr
"""

from openpyxl import load_workbook

wbpath = input("Please input the location of the excel sheet: ")

wb = load_workbook(filename = wbpath, data_only=True)

rowToRead = 2
sheetNumToRead = 0
sheet = wb.worksheets[sheetNumToRead]

gen = sheet.iter_rows(1,2, values_only=True)

iteration = 0
keys = None
values = None

output = dict()
for cells in gen:
    if iteration == 0:
        keys = cells
    else:
        values = cells
    iteration+=1
for i in range(len(keys)):
    output[keys[i]] = values[i]
    
print("output dictionary:")
print(output)